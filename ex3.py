import csv
import xlsxwriter
import openpyxl

#Nombre del archivo CSV
fileCSV = "file.csv"
#Nombre dle archivo xlsx
fileXLSX = "fileExcel.xlsx"
#Tipos de datos
dataType = ["Name", "Surname", "Age"]
#Insertamos los datos
data = [["Ivan", "Petrov", "20"], ["Sergio", "Ramos", "40"]]

def escriureCSV():

    #Escribimos en un archivo csv
    with open(fileCSV, "w") as f:
        #Creamos una variable para escribir los datos
        thewriter = csv.writer(f)

        #Insertamos los tipos de datos
        thewriter.writerow(dataType)

        #Insertamos los datos
        thewriter.writerows(data)
        # thewriter.writerow(["Ivan", "Petrov", "20"])
        # thewriter.writerow(["Sergio", "Ramos", "40"])

#Escribimos el archivo CSV
escriureCSV()

def llegirCSV():

    #Leemos el archivo
    with open(fileCSV) as f:
        # Creamos una variable para leer los datos
        reader = csv.reader(f)

        #hacemos un bucle para recorrerlo
        for row in reader:
            print(row)

#Mostramos el archivo CSV por pantalla
llegirCSV()


def escriureXLSX():
    row = 0
    col = 0
 
    #Especificamos el archivo XLSX
    workbook = xlsxwriter.Workbook(fileXLSX)
    #Creamos el archivo
    worksheet = workbook.add_worksheet()

    #Insertamos los datos
    for name, username, age in (data):
        worksheet.write(row, col, name)
        worksheet.write(row, col + 1, username)
        worksheet.write(row, col + 2, age)
        row += 1

    workbook.close()

#Escribimos el archivo XLSX
escriureXLSX()

def llegirXLSX():
    #Definimos la variable para que cargue los datos del frame
    dataframe = openpyxl.load_workbook(fileXLSX)

    #Definimos la variable para que lea el archivo
    dataframe1 = dataframe.active

    #Recorremos el archivo y mostramos por pantalla los datos
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            print(col[row].value)

#Mostramos el archivo XLSX por pantalla
llegirXLSX()